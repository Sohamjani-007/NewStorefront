from openpyxl import Workbook, load_workbook
import xlsxwriter
import pandas as pd
import openpyxl
import xlrd 
import xlwt


# workbook = load_workbook("test1212.xlsx") # Your Excel file
# worksheet = workbook.active # gets first sheet

# for row in range(1, 10):
#     # Writes a new value PRESERVING cell styles.
#     worksheet.cell(row=row, column=10, value=f' {row * 10}')

# workbook.save("test1212.xlsx")


# from openpyxl import Workbook
# import time

# book = Workbook()
# sheet = book.active

# sheet['A1'] = 56
# sheet['A2'] = 43
# sheet['A3'] = 

# book.save("test1212.xlsx")

# wb = load_workbook("test1212.xlsx")
# sheet = wb.worksheets[0]

# numbers = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]


# sheet.cell(row=1, column=10).value="Numbers"

# j=2
# for i in range(0, 5):
#     sheet.cell(row=j, column=1).value=numbers[i]
#     j+=1

# wb.save("test1212.py")


# class ExcelOperations:
#     def read_path(self):
#         path = "./test1212.xlsx"
#         wb_obj = openpyxl.load_workbook(path)
#         sheet_obj = wb_obj.active
#         cell_obj = sheet_obj.cell(row=6, column=3)
#         print(cell_obj.value)
#         print(sheet_obj.max_column)
#         print(sheet_obj.max_row)

#     def read_excel(self):
#         workbook = xlrd.open_workbook("test1212.xlsx")
#         worksheet = workbook.sheet_by_index(0)

#         # Iterate the rows and columns
#         for i in range(0, 5):
#             for j in range(0, 3):
#                 # Print the cell values with tab space
#                 print(worksheet.cell_value(i, j), end='\t')
#             print('')

#     def write_excel(self):
#         wb = xlwt.Workbook()
#         print(wb)
#         sheet1 = wb.add_sheet('Sheet 1')
#         sheet1.write(1, 0, 'ISBT DEHRADUN')
#         sheet1.write(2, 0, 'SHASTRADHARA')
#         sheet1.write(3, 0, 'CLEMEN TOWN')
#         sheet1.write(4, 0, 'RAJPUR ROAD')
#         sheet1.write(5, 0, 'CLOCK TOWER')
#         sheet1.write(6, 1, 'ISBT DEHRADUN')
#         sheet1.write(7, 2, 'SHASTRADHARA')
#         sheet1.write(8, 3, 'CLEMEN TOWN')
#         sheet1.write(9, 4, 'RAJPUR ROAD')
#         sheet1.write(10, 5, 'CLOCK TOWER')

#         wb.save('test1212.xlsx')
#         print(wb)

# excel = ExcelOperations()
# excel.read_path()
# excel.write_excel()



# def read_excel():
#     workbook = xlrd.open_workbook('ExperimentData.xlsx')

#     #Get the first sheet in the workbook by index
#     sheet1 = workbook.sheet_by_index(0)

#     #Get each row in the sheet as a list and print the list
#     for rowNumber in range(sheet1.nrows):
#         row = sheet1.row_values(rowNumber)
#         print(row)


# def write_excel():
#     #Create a workbook object
#     workbook = xlwt.Workbook()

#     #Add a sheet
#     sheet = workbook.add_sheet('sheet 1')

#     #Write values to the sheet by cell number
#     for x in range(1,11):
#         for y in range(1,11):
#             sheet.write(x-1,y-1,x*y)

#     #Save the workbook to the xls format
#     workbook.save('ExperimentData.xlsx')


# import xlrd
# import xlwt
# import numpy as np

# def read_excel(filename, n=0):
#     """Converts first sheet from an Excel file into an ndarray
    
#     Parameters
#     ----------
#     filename : string
#         Path to file.
        
#     Returns
#     -------
#     ndarray with sheet contents (no conversion done)
#     """
#     contentstring = open(filename, 'rb').read()
#     book  = xlrd.open_workbook(file_contents=contentstring)
#     sheet = book.sheets()[n]
#     array = np.zeros((sheet.ncols, sheet.nrows))
    
#     for row in range(sheet.nrows):
#         for col in range(sheet.ncols):
#             array[col][row] = sheet.cell(row, col).value
    
#     return array

# def write_excel(filename, sheetnames, arrays):
#     """Creates an Excel file with given sheet names and arrays.
    
#     Parameters
#     ----------
#     filename : string
#         Path to file.
#     sheetnames : iterable (string)
#         List of names for the book's sheets
#     arrays : iterable (ndarray)
#         List of data arrays for the book's sheets    
#     """
#     if len(sheetnames) != len(arrays):
#         raise IndexError("Array and sheet number must be equal.")
        
#     book = xlwt.Workbook()
    
#     for name, array in zip(sheetnames, arrays):
#         sheet = book.add_sheet(name)
#         cols, rows = array.shape
        
#         for row in range(rows):
#             for col in range(cols):
#                 sheet.write(row, col, array[col][row])
    
#     book.save(filename)

# workbook = xlwt.Workbook()
# sheet = workbook.add_sheet("Marvels")

# listA = ['H1', 'H2', 'H3', 'H4', 'H5', 'H6']
# listB = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
# listC = ['W1', 'W2', 'W3', 'W4', 'W5', 'W6']
# row = 0

# while(row < len(listA)):
#     column = 0
#     sheet.write(row, column, listA[row])
#     column = column + 1
#     sheet.write(row, column, listB[row] * 10)
#     column = column + 1
#     sheet.write(row, column, listC[row])
#     row = row + 1
# # workbook.save("newexcel.xlsx")

# # data = [["1", 10], ["2", 20], ["3", 30]]
# # for i in range(0, 3):
# #     for j in range(0, 2):
# #         sheet.write(i, j, data[i][j])
# workbook.save("newexcel.xlsx")

# import xlrd

# EXCEL_FILES_FOLDER = 'excel_files/'

# excel_file_path = EXCEL_FILES_FOLDER+'read_excel.xlsx'

# loc = (excel_file_path)
# # To open Workbook
# wb = xlrd.open_workbook(loc)

# sheetCust = wb.sheet_by_name('Customer')

# # script to get number of rows and columns in the sheet
# rowsCount = sheetCust.nrows
# colsCount = sheetCust.ncols

# print("rowsCount: ",rowsCount," colsCount: ",colsCount)

# wb = xlrd.open_workbook('newexcel.xlsx')

# # nsheets = wb.nsheets
# # print('Number of sheets:', nsheets)
# class BasicsXl():
#     def __init__(self):
#         pass
#     def read_excel(self):
#         sheets = wb.sheets()
#         nrows, ncols = sheets[0].nrows, sheets[0].ncols


#         first_row = sheets[0].row_values(0)
#         third_column = sheets[0].col_values(2)
#         cell_value_3_3 = sheets[0].cell_value(2, 2)

#         print(first_row)
#         print(third_column)
#         print(cell_value_3_3)

#         print('Number of rows:', nrows)
#         print('Number of columns:', ncols)

#     def write_excel(self):
#         wbwt = xlwt.Workbook(encoding='utf-8')
#         # ws = wbwt.add_sheet('Sheet1', cell_overwrite_ok=True)
#         # wbwt.save('xlwt_save01.xlsx')


#         ws = wbwt.add_sheet('Sheet2', cell_overwrite_ok=True)
#         # ws.write(0, 0, 100)
#         # ws.write(1, 1, 200)
#         # ws.write(2, 2, 300) # 100, 200, 300 are stored in the positions of (0, 0), (1, 1), (2, 2), respectively.
#         # wbwt.save('xlwt_save02.xlsx')

#         for i in range(10):
#             for j in range(10):
#                 ws.write(i, j, (i+j)**2)
#         wbwt.save('xlwt_save03.xlsx')

# basic = BasicsXl()
# basic.write_excel()


# def read_excel(n):
#     workbook = xlsxwriter.Workbook("newsheet3.xlsx")
#     worksheet = workbook.add_worksheet()
#     worksheet.write('A1', 'Header1')
#     worksheet.write('B1', 'H1')
#     worksheet.write('C1', 'H2')
#     worksheet.write('D1', 'H3')
#     worksheet.write('E1', 'H4')
#     worksheet.write('F1', 'H5')
#     worksheet.write('G1', 'H6')
#     worksheet.write('H1', 'H7')
#     worksheet.write('I1', 'H8')
#     worksheet.write('J1', 'H9')
#     worksheet.write('K1', 'H10')
#     row = 1
#     column = 0
#     content = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
#     for item in content:
#         worksheet.write(row, column, item)
        
#         row = row + 1

#     v = n * n
#     num = write_excel(v)
#     return num


# def write_excel(content):
#     workbook = xlsxwriter.Workbook("newsheet3.xlsx")
#     worksheet = workbook.add_worksheet()
#     worksheet.write('A1', 'A1')
#     worksheet.write('B1', 'A2')
#     worksheet.write('C1', 'A3')
#     worksheet.write('D1', 'A4')
#     worksheet.write('E1', 'A5')
#     worksheet.write('F1', 'A6')
#     worksheet.write('G1', 'A7')
#     worksheet.write('H1', 'A8')
#     worksheet.write('I1', 'A9')
#     worksheet.write('J1', 'A10')
#     content = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
#     row = 1
#     column = 0
#     a_list = [item * 10 for item in content]
#     for item in a_list :
#         worksheet.write(row, column, item)      
#         row = row + 1
#     workbook.close()
#     return content
    
# read_excel(10)


class ExcelOperation:
    def read_excel(self):
            workbook = xlrd.open_workbook('/home/oto/Desktop/Project/django-projects/storefront/xlwt_save.xlsx',on_demand=True)
            worksheet = workbook.sheet_by_index(0)
            read_data_list = list()
            for row in range(1, worksheet.nrows):
                h1 = int(worksheet.cell_value(row, 0)) if worksheet.cell_value(row, 0) else 0
                h2 = int(worksheet.cell_value(row, 1)) if worksheet.cell_value(row, 1) else 0
                read_data_row_list = [h1, h2]
                read_data_list.append(read_data_row_list)
            return read_data_list

    def write_excel(self, input_data):
            worksheet = xlwt.Workbook()
            sheet = worksheet.add_sheet('data_multiple_by_ten')
            sheet.write(0, 0, 'A1')
            sheet.write(0, 1, 'A2')
            i = 1
            for input_value in input_data:
                j = 0
                a1 = input_value[0] * 10 if len(input_value) > 0 else 0
                a2 = input_value[1] * 10 if len(input_value) > 1 else 0
                sheet.write(i, j, a1)
                sheet.write(i, j + 1, a2)
                print(a1, a2)
                i+=1
            worksheet.save("data_multiple_by_ten.xlsx")

ops = ExcelOperation()
read_output = ops.read_excel()
ops.write_excel(read_output)



