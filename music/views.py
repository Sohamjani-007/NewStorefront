import os
from django.shortcuts import render
from .models import Contact, Song, Album, Product
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from xlrd import open_workbook  # http://pypi.python.org/pypi/xlrd
import io
import xlrd
import xlwt
import xlsxwriter


class ExcelPageView(TemplateView):
    template_name = "excel_home.html"


# Create your views here.

def showformdata(request):
    t = Contact.objects.get(id=1)
    t.name = "Starlord"
    t.address = "yooo"
    t.save()

    return render(request, 'userregistration.html', dict())


def product_data(request):
    file_path = '/home/oto/Desktop/Project/django-projects/storefront/request_file.xlsx'
    wb = xlrd.open_workbook(file_path)
    wb_sheet = wb.sheet_by_index(0)
    total_rows = wb_sheet.nrows
    print(total_rows)

    for rownum in range(1, total_rows):
        product_name = wb_sheet.cell(rownum, 0).value
        product_price = wb_sheet.cell(rownum, 1).value
        print(product_name, product_price)
        if product_name and product_price:
            Product.objects.filter(product_name=product_name).update(product_price=product_price)
            g = Product.objects.get(id=4)
            g.product_name = "Ninja"
            g.product_price = "100"
            g.save()
    return render(request, 'product_registration.html', dict())


def export_xlsx(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("MyModel")

    row_num = 0

    columns = [
        (u"ID", 2000),
        (u"Title", 6000),
        (u"Description", 8000),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    read_data_list = list()
    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.title,
            obj.description,
        ]

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


export_xlsx.short_description = u"Export XLSX"


class ExcelOperation:
    def read_excel(self):
        try:
            workbook = xlrd.open_workbook(
                '/home/oto/Desktop/Project/django-projects/storefront/xlwt_save01.xlsx', on_demand=True)
            worksheet = workbook.sheet_by_index(0)
            read_data_list = list()
            for row in range(1, worksheet.nrows):
                h1 = int(worksheet.cell_value(row, 0)
                         ) if worksheet.cell_value(row, 0) else 0
                h2 = int(worksheet.cell_value(row, 1)
                         ) if worksheet.cell_value(row, 1) else 0
                read_data_row_list = [h1, h2]
                read_data_list.append(read_data_row_list)
            return read_data_list
        except Exception as e:
            print(e.__str__())
            pass

    def write_excel(self, input_data):
        try:

            worksheet = xlwt.Workbook()
            sheet = worksheet.add_sheet('data_multiple_by_ten')
            sheet.write(0, 0, 'Yo')
            sheet.write(0, 1, 'Howdy')
            i = 1
            for input_value in input_data:
                j = 0
                a1 = input_value[0] * 10 if len(input_value) > 0 else 0
                a2 = input_value[1] * 10 if len(input_value) > 1 else 0
                sheet.write(i, j, a1)
                sheet.write(i, j + 1, a2)
                print(a1, a2)
                i += 1
            worksheet.save("data_multiple_by_ten.xlsx")
        except Exception as e:
            print(e.__str__())
            pass


# ops = ExcelOperation()
# read_output = ops.read_excel()
# ops.write_excel(read_output)


# def xlsx(request):
#     # Create a workbook and add a worksheet.
#     output = io.BytesIO()
#     workbook = xlsxwriter.Workbook(output, {'in_memory': True})
#     worksheet = workbook.add_worksheet('Reporte3a5')
#     bold = workbook.add_format({'bold': True})
#     # Some data we want to write to the worksheet.
#     reporte = Album.objects.all() #my model

#     # Start from the first cell. Rows and columns are zero indexed.
#     row = 1
#     col = 0

#     # Iterate over the data and write it out row by row.
#     for linea in reporte:
#         worksheet.write(row, col, linea.artist)
#         worksheet.write(row, col + 1, linea.album_title)
#         worksheet.write(row, col + 2, linea.genre)
#         row += 1

#     # Write the title for every column in bold
#     worksheet.write('A1', 'Priority', bold)
#     worksheet.write('B1', 'Code', bold)

#     workbook.close()

#     output.seek(0)
#     response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = "attachment; filename=Reporte3a5.xlsx"

#     return response


# def export_abc_xlsx(request):
#     response = HttpResponse(content_type='application/ms-excel')
#     response['Content-Disposition'] = 'attachment; filename="abc.xlsx"'

#     wb = xlwt.Workbook(encoding='utf-8')
#     ws = wb.add_sheet('Users Data')

#     row_num = 0
#     columns = ['Username', 'First Name', 'Last Name', 'Email Address', ]
#     for col_num in range(len(columns)):
#         ws.write(row_num, col_num, columns[col_num])


#     rows = Product.objects.all().values_list('product_name', 'product_price')
#     for row in rows:
#         row_num += 1
#         for col_num in range(len(row)):
#             ws.write(row_num, col_num, row[col_num])
#     wb.save(response)
#     return render(request, response,  dict())


# def export_write_xls(request):
#     response = HttpResponse(content_type='application/ms-excel')
#     response['Content-Disposition'] = 'attachment; filename="users.xls"'

#     path = os.path.dirname(__file__)
#     template_name = "excel_home.html"
#     file = os.path.join(path, 'sample.xls')

#     rb = open_workbook(file, formatting_info=True)
#     r_sheet = rb.sheet_by_index(0)

#     wb = copy(rb)
#     ws = wb.get_sheet(0)

#     row_num = 2 # index start from 0
#     rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
#     for row in rows:
#         row_num += 1
#         for col_num in range(len(row)):
#             ws.write(row_num, col_num, row[col_num])

#     # wb.save(file) # will replace original file
#     # wb.save(file + '.out' + os.path.splitext(file)[-1]) # will save file where the excel file is
#     wb.save(response)
#     return response


# def export(request):
#     person_resource = Product()
#     dataset = person_resource.export()
#     response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = 'attachment; filename="persons.xls"'
#     return response

# def simple_upload(request):
#     if request.method == 'POST':
#         person_resource = PersonResource()
#         dataset = Dataset()
#         new_persons = request.FILES['myfile']

#         imported_data = dataset.load(new_persons.read(),format='xlsx')
#         #print(imported_data)
#         for data in imported_data:
#         	print(data[1])
#         	value = Person(
#         		data[0],
#         		data[1],
#         		 data[2],
#         		 data[3]
#         		)
#         	value.save()

        # result = person_resource.import_data(dataset, dry_run=True)  # Test the data import

        # if not result.has_errors():
        #    person_resource.import_data(dataset, dry_run=False)  # Actually import now

    # return render(request, 'input.html')

# def export_excel(request):
#     response = HttpResponse(content_type='application/ms-excel')
#     response['Content-Disposition'] = f'attachment; filename="filename.xlsx"'

#     wb = Workbook()
#     wb.save(response)
#     return response
